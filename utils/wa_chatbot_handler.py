"""
utils/wa_chatbot_handler.py
Handler chatbot – WA & Web Widget.
Menu:
  1. Cek tagihan
  2. Info kamar (jumlah, kosong/isi dari DB, foto/video)
  3. Foto & video fasilitas
  4. Harga kamar (dari Excel detailkamar.xlsx)
  5. Fasilitas kost (dari Excel detailkamar.xlsx)
  + Hubungi admin
"""
import os
import time
import json
import re
from models.database import get_db

# ── Path Excel ────────────────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH   = os.path.join(BASE_DIR, 'detailkamar.xlsx')

# ── Folder media ──────────────────────────────────────────────────────────────
MEDIA_KAMAR_DIR     = os.path.join(BASE_DIR, 'static', 'media_kamar')
MEDIA_FASILITAS_DIR = os.path.join(BASE_DIR, 'static', 'media_fasilitas')
FOTO_EXT  = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}
VIDEO_EXT = {'.mp4', '.webm', '.mov'}

ADMIN_NO = "08159959605"
TOTAL_KAMAR = 33

# Emoji angka → plain digit
_EMOJI_NUM = {
    '1️⃣': '1', '2️⃣': '2', '3️⃣': '3',
    '4️⃣': '4', '5️⃣': '5', '6️⃣': '6',
    '7️⃣': '7', '8️⃣': '8', '9️⃣': '9', '0️⃣': '0',
}

def _normalize(msg: str) -> str:
    """Strip whitespace dan convert emoji angka ke digit biasa."""
    msg = msg.strip()
    for emoji, digit in _EMOJI_NUM.items():
        msg = msg.replace(emoji, digit)
    return msg.strip().lower()

# ── Session ───────────────────────────────────────────────────────────────────
_wa_state:     dict = {}
_wa_last_seen: dict = {}
_web_state:    dict = {}
_web_last_seen:dict = {}
_SESSION_TIMEOUT = 3600   # 1 jam

# ── Cache Excel (TTL 5 menit) ─────────────────────────────────────────────────
_excel_cache:      dict = {}
_excel_cache_time: float = 0
_EXCEL_TTL = 300

# ═════════════════════════════════════════════════════════════════════════════
# LOAD DATA EXCEL
# ═════════════════════════════════════════════════════════════════════════════

def _load_excel() -> dict:
    """Baca detailkamar.xlsx, kembalikan dict berisi data kamar, fasilitas, harga, info."""
    global _excel_cache, _excel_cache_time
    now = time.time()
    if _excel_cache and (now - _excel_cache_time) < _EXCEL_TTL:
        return _excel_cache

    data = {
        'kamar': [],        # list dict dari sheet "Info Kamar"
        'fasilitas': [],    # list dict dari sheet "Fasilitas Umum"
        'harga': [],        # list dict dari sheet "Harga & Paket"
        'info': {},         # dict key->value dari sheet "Info Kost"
    }

    if not os.path.isfile(EXCEL_PATH):
        return data

    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

        # Sheet Info Kamar
        if 'Info Kamar' in wb.sheetnames:
            ws = wb['Info Kamar']
            cols = None
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0] == 'No. Kamar *' or (cols is None and str(row[0]).startswith('No.')):
                    # Normalisasi header: gabungkan baris baru jadi spasi tunggal
                    cols = [' '.join(str(c or '').split()) for c in row]
                    continue
                if cols and row[0]:
                    vals = [' '.join(str(v or '').split()) for v in row]
                    item = dict(zip(cols, vals))
                    # Coba beberapa variasi header untuk kolom yang mungkin punya newline
                    def _get_col(item, *keys, default=''):
                        for k in keys:
                            if k in item and item[k]:
                                return item[k]
                        return default
                    d = {
                        'nomor':     _get_col(item, 'No. Kamar *'),
                        'lantai':    _get_col(item, 'Lantai'),
                        'tipe':      _get_col(item, 'Tipe Kamar *'),
                        'harga':     _get_col(item, 'Harga/Bulan (Rp) *', default='0'),
                        'status':    _get_col(item, 'Status * (kosong/isi)', 'Status *(kosong/isi)',
                                              'Status *', 'Status', default='isi').lower(),
                        'luas':      _get_col(item, 'Luas (m²)'),
                        'kapasitas': _get_col(item, 'Kapasitas (orang)', 'Kapasitas'),
                        'deskripsi': _get_col(item, 'Deskripsi Kamar'),
                        'fasilitas': _get_col(item, 'Fasilitas Dalam Kamar'),
                        'catatan':   _get_col(item, 'Catatan Tambahan'),
                    }
                    if d['nomor']:
                        data['kamar'].append(d)

        # Sheet Fasilitas Umum
        if 'Fasilitas Umum' in wb.sheetnames:
            ws = wb['Fasilitas Umum']
            cols = None
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0] and 'Nama Fasilitas' in str(row[0]):
                    cols = [' '.join(str(c or '').split()) for c in row]
                    continue
                if cols and row[0]:
                    vals = [' '.join(str(v or '').split()) for v in row]
                    item = dict(zip(cols, vals))
                    d = {
                        'nama':      item.get('Nama Fasilitas *', ''),
                        'kategori':  item.get('Kategori', ''),
                        'kondisi':   (item.get('Kondisi (baik/rusak/perlu perbaikan)', '') or
                                      item.get('Kondisi', '') or 'baik'),
                        'jumlah':    item.get('Jumlah', ''),
                        'lokasi':    item.get('Lokasi di Kost', ''),
                        'keterangan':item.get('Keterangan Tambahan', ''),
                    }
                    if d['nama'] and not d['nama'].startswith('📌'):
                        data['fasilitas'].append(d)

        # Sheet Harga & Paket
        if 'Harga & Paket' in wb.sheetnames:
            ws = wb['Harga & Paket']
            cols = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and 'Tipe Kamar' in str(row[0]):
                    cols = [str(c or '').strip() for c in row]
                    continue
                if cols and row[0] and str(row[0]).strip() not in ('', '📋  BIAYA TAMBAHAN & ATURAN'):
                    item = dict(zip(cols, row))
                    tipe = str(item.get('Tipe Kamar *', '') or '').strip()
                    if tipe and not tipe.startswith(('📋', 'Listrik', 'Air', 'WiFi', 'Laundry', 'Kebersihan', 'Denda')):
                        try:
                            harga_bln = int(float(str(item.get('Harga Bulanan (Rp) *', 0) or 0)))
                        except Exception:
                            harga_bln = 0
                        try:
                            deposit = int(float(str(item.get('Deposit (Rp)', 0) or 0)))
                        except Exception:
                            deposit = 0
                        data['harga'].append({
                            'tipe':    tipe,
                            'bulanan': harga_bln,
                            'deposit': deposit,
                            'ket':     str(item.get('Keterangan', '') or ''),
                        })

        # Sheet Info Kost
        if 'Info Kost' in wb.sheetnames:
            ws = wb['Info Kost']
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0] and row[1]:
                    key = str(row[0]).strip()
                    val = str(row[1]).strip()
                    if key not in ('Keterangan',):
                        data['info'][key] = val

        wb.close()

    except Exception as e:
        print(f"[Excel] Gagal baca: {e}")

    _excel_cache      = data
    _excel_cache_time = now
    return data


# ═════════════════════════════════════════════════════════════════════════════
# HELPER MEDIA
# ═════════════════════════════════════════════════════════════════════════════

def _list_media(folder: str) -> dict:
    """Kembalikan {'foto': [...url], 'video': [...url]}"""
    result = {'foto': [], 'video': []}
    if not os.path.isdir(folder):
        return result
    for fname in sorted(os.listdir(folder)):
        ext = os.path.splitext(fname)[1].lower()
        if ext in FOTO_EXT:
            result['foto'].append(fname)
        elif ext in VIDEO_EXT:
            result['video'].append(fname)
    return result


def _media_url_kamar(fname: str) -> str:
    return f"/media-kamar/file/{fname}"


def _media_url_fasilitas(fname: str) -> str:
    return f"/media-fasilitas/file/{fname}"


# ═════════════════════════════════════════════════════════════════════════════
# DB HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def _kamar_dari_db():
    """Kembalikan dict {nomor_kamar: {'aktif': 0/1, 'harga': x}} dari DB."""
    try:
        conn = get_db()
        rows = conn.execute(
            "SELECT nomor_kamar, aktif, harga_sewa FROM penghuni ORDER BY nomor_kamar"
        ).fetchall()
        conn.close()
        return {r['nomor_kamar']: {'aktif': r['aktif'], 'harga': r['harga_sewa']} for r in rows}
    except Exception:
        return {}


def _get_penghuni_by_phone(sender: str):
    try:
        conn   = get_db()
        suffix = sender[-9:]
        row    = conn.execute(
            """SELECT id, nama, nomor_kamar, harga_sewa
               FROM penghuni
               WHERE REPLACE(REPLACE(REPLACE(no_hp,' ',''),'-',''),'+','') LIKE ?
               AND aktif = 1""",
            (f'%{suffix}',)
        ).fetchone()
        conn.close()
        return dict(row) if row else None
    except Exception:
        return None


def _format_bulan(s: str) -> str:
    bln = ['','Jan','Feb','Mar','Apr','Mei','Jun','Jul','Agu','Sep','Okt','Nov','Des']
    try:
        y, m = s.split('-')
        return f"{bln[int(m)]} {y}"
    except Exception:
        return s


def _fmt_rp(n) -> str:
    try:
        return f"Rp {int(float(str(n))):,}".replace(',', '.')
    except Exception:
        return str(n)


# ═════════════════════════════════════════════════════════════════════════════
# WA HANDLER (plain text)
# ═════════════════════════════════════════════════════════════════════════════

MENU_WA = f"""Halo kak! 👋 Selamat datang di *Kost Kami*.

Ketik angka untuk pilih menu:
1️⃣  Cek tagihan saya
2️⃣  Info kamar (kosong/isi + foto)
3️⃣  Foto & video fasilitas
4️⃣  Harga kamar
5️⃣  Fasilitas kost
6️⃣  Hubungi admin

Ketik *menu* kapan saja untuk kembali."""


def handle_incoming_wa(sender: str, message_text: str) -> str:
    message_text = message_text.strip()
    if not message_text:
        return ''

    _cleanup(_wa_state, _wa_last_seen)
    _wa_last_seen[sender] = time.time()
    if sender not in _wa_state:
        _wa_state[sender] = {'step': 'main'}

    msg   = _normalize(message_text)
    state = _wa_state[sender]

    if msg in ('menu', 'halo', 'hai', 'hi', 'hello', 'help', 'bantuan', '0'):
        _wa_state[sender] = {'step': 'main', 'unknown_count': 0}
        return MENU_WA

    if state['step'] == 'main':
        reply = _wa_main(sender, msg)
    elif state['step'] == 'cek_tagihan_input':
        reply = _wa_cek_tagihan(sender, message_text)
    else:
        _wa_state[sender] = {'step': 'main'}
        reply = MENU_WA

    try:
        _log_wa(sender, message_text, reply)
    except Exception as e:
        print(f"[WA Handler] _log_wa error (diabaikan): {e}")
    return reply


def _wa_main(sender: str, msg: str) -> str:
    penghuni = _get_penghuni_by_phone(sender)

    # ── Cek angka murni DULU sebelum keyword match ──────────────────────
    if msg == '1':
        _wa_state[sender]['unknown_count'] = 0
        if penghuni:
            return _tagihan_teks(penghuni['id'], penghuni['nama'])
        _wa_state[sender] = {'step': 'cek_tagihan_input'}
        return "Silakan ketik *nama* atau *nomor kamar* kamu kak.\n_(Ketik *menu* untuk kembali)_"
    elif msg == '2':
        _wa_state[sender]['unknown_count'] = 0
        return _wa_info_kamar()
    elif msg == '3':
        _wa_state[sender]['unknown_count'] = 0
        return _wa_media_fasilitas()
    elif msg == '4':
        _wa_state[sender]['unknown_count'] = 0
        return _wa_harga()
    elif msg == '5':
        _wa_state[sender]['unknown_count'] = 0
        return _wa_fasilitas()
    elif msg == '6':
        _wa_state[sender]['unknown_count'] = 0
        return _wa_admin()

    # ── Keyword match (teks bebas) ─────────────────────────────────────
    elif any(k in msg for k in ('tagihan','bayar','cicil','tunggak','lunas')):
        _wa_state[sender]['unknown_count'] = 0
        if penghuni:
            return _tagihan_teks(penghuni['id'], penghuni['nama'])
        _wa_state[sender] = {'step': 'cek_tagihan_input'}
        return "Silakan ketik *nama* atau *nomor kamar* kamu kak.\n_(Ketik *menu* untuk kembali)_"
    elif any(k in msg for k in ('kamar kosong','kamar tersedia','sewa kamar','foto kamar','video kamar','info kamar')):
        return _wa_info_kamar()
    elif any(k in msg for k in ('foto fasilitas','video fasilitas','galeri fasilitas','media fasilitas')):
        return _wa_media_fasilitas()
    elif any(k in msg for k in ('harga','biaya','tarif','berapa')):
        return _wa_harga()
    elif any(k in msg for k in ('fasilitas','lengkap','apa saja','fitur')):
        return _wa_fasilitas()
    elif any(k in msg for k in ('admin','hubungi','kontak','cs','telpon')):
        return _wa_admin()

    else:
        # Hitung berapa kali berturut-turut tidak dikenali
        unk = state.get('unknown_count', 0) + 1
        _wa_state[sender]['unknown_count'] = unk
        if unk >= 2:
            _wa_state[sender] = {'step': 'main'}  # reset
            return (
                f"Maaf kak, kami tidak bisa memahami pesanmu. 🙏\n"
                f"Menghubungkan ke admin ya...\n\n"
                f"📞 Silakan chat langsung: *{ADMIN_NO}*"
            )
        return f"Maaf kak, perintah tidak dikenali 🙏\n\n{MENU_WA}"


def _wa_cek_tagihan(sender: str, input_text: str) -> str:
    _wa_state[sender] = {'step': 'main'}
    conn = get_db()
    row  = conn.execute(
        "SELECT id, nama FROM penghuni WHERE LOWER(nomor_kamar)=? AND aktif=1",
        (input_text.lower(),)
    ).fetchone()
    if not row:
        row = conn.execute(
            "SELECT id, nama FROM penghuni WHERE LOWER(nama) LIKE ? AND aktif=1",
            (f'%{input_text.lower()}%',)
        ).fetchone()
    conn.close()
    if not row:
        return f"❌ Penghuni *{input_text}* tidak ditemukan.\nHubungi admin: {ADMIN_NO}"
    return _tagihan_teks(row['id'], row['nama'])


def _tagihan_teks(pid: int, nama: str) -> str:
    conn = get_db()
    rows = conn.execute(
        "SELECT bulan, jumlah, status FROM tagihan WHERE penghuni_id=? ORDER BY bulan DESC LIMIT 6",
        (pid,)
    ).fetchall()
    conn.close()
    if not rows:
        return f"Halo kak *{nama}* 😊\nBelum ada data tagihan."
    lines = [f"📋 *Tagihan – {nama}*\n"]
    for r in rows:
        ico = '✅' if r['status'] == 'lunas' else '⏳' if r['status'] == 'cicilan' else '❌'
        lines.append(f"{ico} {_format_bulan(r['bulan'])}: {_fmt_rp(r['jumlah'])} – _{r['status']}_")
    belum = [r for r in rows if r['status'] != 'lunas']
    if belum:
        lines.append(f"\n💰 Belum lunas: *{_fmt_rp(sum(r['jumlah'] for r in belum))}*")
        lines.append(f"📞 Hubungi: {ADMIN_NO}")
    else:
        lines.append("\n✅ Semua lunas. Terima kasih! 🙏")
    return '\n'.join(lines)


def _wa_info_kamar() -> str:
    # Data utama dari DB penghuni
    db_data = _kamar_dari_db()
    terisi  = sum(1 for v in db_data.values() if v['aktif'] == 1)
    kosong  = TOTAL_KAMAR - terisi

    # Kamar kosong (aktif=0 di DB atau belum terdaftar di DB)
    kamar_kosong_db = [k for k, v in db_data.items() if v['aktif'] == 0]

    lines = [f"🏠 *Info Kamar Kost Kami*\n"]
    lines.append(f"📊 Total: {TOTAL_KAMAR} kamar | Terisi: {terisi} | *Kosong: {kosong}*\n")

    if kamar_kosong_db:
        lines.append("*Kamar Tersedia:*")
        for nomor in sorted(kamar_kosong_db):
            harga = db_data[nomor]['harga']
            lines.append(f"  • Kamar {nomor}: {_fmt_rp(harga)}/bulan")
    else:
        lines.append("😔 Saat ini semua kamar *penuh*.")

    # Info foto
    media = _list_media(MEDIA_KAMAR_DIR)
    jml_foto  = len(media['foto'])
    jml_video = len(media['video'])
    if jml_foto or jml_video:
        lines.append(f"\n📸 Tersedia {jml_foto} foto & {jml_video} video kamar.")
        lines.append("👉 Lihat galeri di website kost kami atau minta link ke admin.")

    lines.append(f"\n📞 Info & booking: *{ADMIN_NO}*")
    return '\n'.join(lines)


def _wa_media_fasilitas() -> str:
    media = _list_media(MEDIA_FASILITAS_DIR)
    jml_foto  = len(media['foto'])
    jml_video = len(media['video'])
    if not jml_foto and not jml_video:
        return (
            "📸 *Foto & Video Fasilitas*\n\n"
            "Maaf, foto/video fasilitas belum tersedia saat ini.\n"
            f"Untuk melihat langsung, hubungi admin: *{ADMIN_NO}*"
        )
    lines = [f"📸 *Foto & Video Fasilitas Kost Kami*\n"]
    lines.append(f"• 🖼️ {jml_foto} foto fasilitas tersedia")
    lines.append(f"• 🎥 {jml_video} video fasilitas tersedia")
    lines.append("\n👉 Lihat galeri lengkap di website kost kami.")
    lines.append(f"📞 Atau minta link ke admin: *{ADMIN_NO}*")
    return '\n'.join(lines)


def _wa_harga() -> str:
    xdata = _load_excel()
    lines = ["💰 *Harga Kamar Kost Kami*\n"]
    if xdata['harga']:
        for h in xdata['harga']:
            lines.append(f"🏠 *{h['tipe']}*")
            lines.append(f"   • Bulanan: {_fmt_rp(h['bulanan'])}/bulan")
            if h['deposit']:
                lines.append(f"   • Deposit: {_fmt_rp(h['deposit'])}")
            if h['ket']:
                lines.append(f"   _{h['ket']}_")
            lines.append("")
    else:
        # Fallback dari DB
        db_data = _kamar_dari_db()
        harga_set = sorted(set(v['harga'] for v in db_data.values() if v['harga'] > 0))
        if harga_set:
            for h in harga_set:
                lines.append(f"   • {_fmt_rp(h)}/bulan")
        else:
            lines.append("Hubungi admin untuk info harga terkini.")

    lines.append(f"📞 Info lebih lanjut: *{ADMIN_NO}*")
    return '\n'.join(lines)


def _wa_fasilitas() -> str:
    xdata = _load_excel()
    lines = ["✨ *Fasilitas Kost Kami*\n"]
    if xdata['fasilitas']:
        by_kat: dict = {}
        for f in xdata['fasilitas']:
            kat = f['kategori'] or 'Umum'
            by_kat.setdefault(kat, []).append(f)
        for kat, items in by_kat.items():
            lines.append(f"*{kat}:*")
            for item in items:
                kondisi_ico = '✅' if 'baik' in item['kondisi'].lower() else '⚠️'
                ket = f" – {item['keterangan']}" if item['keterangan'] else ''
                lines.append(f"  {kondisi_ico} {item['nama']}{ket}")
            lines.append("")
    else:
        lines.append(
            "• WiFi gratis\n• Parkir motor & mobil\n• Dapur bersama\n"
            "• Mesin cuci\n• CCTV 24 jam\n• Keamanan 24 jam"
        )
    lines.append(f"📞 Info: *{ADMIN_NO}*")
    return '\n'.join(lines)


def _wa_admin() -> str:
    xdata  = _load_excel()
    nama   = xdata['info'].get('Nama Admin / Pemilik', '')
    jam    = xdata['info'].get('Jam Layanan Admin', '08.00 – 21.00 WIB')
    no_hp  = xdata['info'].get('No. HP / WA Admin', ADMIN_NO)
    return (
        f"📞 *Hubungi Admin Kost Kami*\n\n"
        f"{'👤 ' + nama + chr(10) if nama else ''}"
        f"WA / Telp: *{no_hp}*\n"
        f"Jam: {jam}\n\n"
        f"Silakan hubungi langsung ya kak 🙏"
    )


# ═════════════════════════════════════════════════════════════════════════════
# WEB WIDGET HANDLER (HTML)
# ═════════════════════════════════════════════════════════════════════════════

MENU_WEB = """Halo kak! 👋 Selamat datang di <b>Kost Kami</b>.

Silakan pilih menu di bawah atau ketik pertanyaan:<br><br>
<span class='menu-item' onclick='sendQuick("1")'>1️⃣ Cek tagihan saya</span><br>
<span class='menu-item' onclick='sendQuick("2")'>2️⃣ Info kamar (kosong/isi)</span><br>
<span class='menu-item' onclick='sendQuick("3")'>3️⃣ Foto & video fasilitas</span><br>
<span class='menu-item' onclick='sendQuick("4")'>4️⃣ Harga kamar</span><br>
<span class='menu-item' onclick='sendQuick("5")'>5️⃣ Fasilitas kost</span><br>
<span class='menu-item' onclick='sendQuick("6")'>6️⃣ Hubungi admin</span><br><br>
<i>Ketik "menu" kapan saja untuk kembali.</i>"""


def handle_web_message(session_id: str, message_text: str) -> str:
    message_text = message_text.strip()
    if not message_text:
        return ''

    _cleanup(_web_state, _web_last_seen)
    _web_last_seen[session_id] = time.time()
    if session_id not in _web_state:
        _web_state[session_id] = {'step': 'main'}

    msg   = _normalize(message_text)
    state = _web_state[session_id]

    if msg in ('menu', 'halo', 'hai', 'hi', 'hello', 'help', '0', 'bantuan'):
        _web_state[session_id] = {'step': 'main', 'unknown_count': 0}
        return MENU_WEB

    if state['step'] == 'cek_tagihan_input':
        _web_state[session_id] = {'step': 'main'}
        return _web_cek_tagihan(message_text)

    # ── Cek angka murni DULU sebelum keyword match ──────────────────────
    if msg == '1':
        _web_state[session_id]['unknown_count'] = 0
        _web_state[session_id] = {'step': 'cek_tagihan_input'}
        return "Silakan ketik <b>nama</b> atau <b>nomor kamar</b> kamu kak.<br><i>(contoh: Budi atau A1)</i>"
    elif msg == '2':
        _web_state[session_id]['unknown_count'] = 0
        return _web_info_kamar()
    elif msg == '3':
        _web_state[session_id]['unknown_count'] = 0
        return _web_media_fasilitas()
    elif msg == '4':
        _web_state[session_id]['unknown_count'] = 0
        return _web_harga()
    elif msg == '5':
        _web_state[session_id]['unknown_count'] = 0
        return _web_fasilitas()
    elif msg == '6':
        _web_state[session_id]['unknown_count'] = 0
        return _web_admin()

    # ── Keyword match (teks bebas) ─────────────────────────────────────
    elif any(k in msg for k in ('tagihan','bayar','tunggak','lunas')):
        _web_state[session_id] = {'step': 'cek_tagihan_input'}
        return "Silakan ketik <b>nama</b> atau <b>nomor kamar</b> kamu kak.<br><i>(contoh: Budi atau A1)</i>"
    elif any(k in msg for k in ('kamar kosong','kamar tersedia','sewa kamar','foto kamar','video kamar','info kamar')):
        return _web_info_kamar()
    elif any(k in msg for k in ('foto fasilitas','video fasilitas','galeri fasilitas','media fasilitas')):
        return _web_media_fasilitas()
    elif any(k in msg for k in ('harga','biaya','tarif','berapa','sewa')):
        return _web_harga()
    elif any(k in msg for k in ('fasilitas','lengkap','apa saja','fitur')):
        return _web_fasilitas()
    elif any(k in msg for k in ('admin','hubungi','kontak','cs')):
        return _web_admin()

    else:
        unk = _web_state[session_id].get('unknown_count', 0) + 1
        _web_state[session_id]['unknown_count'] = unk
        if unk >= 2:
            _web_state[session_id] = {'step': 'main'}  # reset
            return (
                f"Maaf kak, kami tidak bisa memahami pesanmu. 🙏<br>"
                f"Menghubungkan ke admin ya...<br><br>"
                f"📞 Chat langsung admin: <b>{ADMIN_NO}</b>"
            )
        return f"Maaf kak, tidak dikenali 🙏<br><br>{MENU_WEB}"


def _web_cek_tagihan(input_text: str) -> str:
    conn = get_db()
    row  = conn.execute(
        "SELECT id, nama FROM penghuni WHERE LOWER(nomor_kamar)=? AND aktif=1",
        (input_text.lower(),)
    ).fetchone()
    if not row:
        row = conn.execute(
            "SELECT id, nama FROM penghuni WHERE LOWER(nama) LIKE ? AND aktif=1",
            (f'%{input_text.lower()}%',)
        ).fetchone()
    conn.close()
    if not row:
        return f"❌ Penghuni <b>{input_text}</b> tidak ditemukan.<br>Hubungi admin: <b>{ADMIN_NO}</b>"

    conn = get_db()
    rows = conn.execute(
        "SELECT bulan, jumlah, status FROM tagihan WHERE penghuni_id=? ORDER BY bulan DESC LIMIT 6",
        (row['id'],)
    ).fetchall()
    conn.close()
    if not rows:
        return f"Halo kak <b>{row['nama']}</b> 😊<br>Belum ada data tagihan."

    lines = [f"📋 <b>Tagihan – {row['nama']}</b><br>"]
    for r in rows:
        ico = '✅' if r['status'] == 'lunas' else '❌'
        lines.append(f"{ico} {_format_bulan(r['bulan'])}: {_fmt_rp(r['jumlah'])} – <i>{r['status']}</i>")
    belum = [r for r in rows if r['status'] != 'lunas']
    if belum:
        lines.append(f"<br>💰 Belum lunas: <b>{_fmt_rp(sum(r['jumlah'] for r in belum))}</b>")
        lines.append(f"📞 Hubungi: {ADMIN_NO}")
    else:
        lines.append("<br>✅ Semua <b>lunas</b>. Terima kasih! 🙏")
    return '<br>'.join(lines)


def _web_info_kamar() -> str:
    db_data = _kamar_dari_db()
    terisi  = sum(1 for v in db_data.values() if v['aktif'] == 1)
    kosong  = TOTAL_KAMAR - terisi
    kamar_kosong = [k for k, v in db_data.items() if v['aktif'] == 0]

    lines = ["🏠 <b>Info Kamar Kost Kami</b><br>"]
    lines.append(f"📊 Total: <b>{TOTAL_KAMAR}</b> kamar | Terisi: {terisi} | Kosong: <b style='color:#27ae60'>{kosong}</b><br>")

    if kamar_kosong:
        lines.append("<b>Kamar Tersedia:</b>")
        for nomor in sorted(kamar_kosong):
            harga = db_data[nomor]['harga']
            lines.append(f"&nbsp;&nbsp;• Kamar {nomor}: {_fmt_rp(harga)}/bulan")
    else:
        lines.append("😔 Saat ini semua kamar <b>penuh</b>.")

    # Foto kamar
    media = _list_media(MEDIA_KAMAR_DIR)
    if media['foto'] or media['video']:
        lines.append(f"<br>📸 Ada <b>{len(media['foto'])}</b> foto & <b>{len(media['video'])}</b> video kamar.")
        # Tampilkan 3 thumbnail foto pertama
        thumbs = media['foto'][:3]
        if thumbs:
            imgs = ' '.join(
                f'<a href="{_media_url_kamar(f)}" target="_blank">'
                f'<img src="{_media_url_kamar(f)}" style="width:90px;height:68px;object-fit:cover;border-radius:6px;margin:2px"></a>'
                for f in thumbs
            )
            lines.append(f"<br>{imgs}")
        if len(media['foto']) > 3:
            lines.append(f'<a href="/media-kamar/list" target="_blank">Lihat semua foto →</a>')

    lines.append(f"<br>📞 Info & booking: <b>{ADMIN_NO}</b>")
    return '<br>'.join(lines)


def _web_media_fasilitas() -> str:
    media = _list_media(MEDIA_FASILITAS_DIR)
    if not media['foto'] and not media['video']:
        return (
            "📸 <b>Foto & Video Fasilitas</b><br><br>"
            "Maaf, media fasilitas belum tersedia saat ini.<br>"
            f"Hubungi admin untuk info lebih lanjut: <b>{ADMIN_NO}</b>"
        )
    lines = ["📸 <b>Foto & Video Fasilitas Kost Kami</b><br>"]
    thumbs = media['foto'][:4]
    if thumbs:
        imgs = ' '.join(
            f'<a href="{_media_url_fasilitas(f)}" target="_blank">'
            f'<img src="{_media_url_fasilitas(f)}" style="width:90px;height:68px;object-fit:cover;border-radius:6px;margin:2px"></a>'
            for f in thumbs
        )
        lines.append(imgs)
    if media['video']:
        lines.append(f"<br>🎥 <b>{len(media['video'])}</b> video fasilitas tersedia.")
    if len(media['foto']) > 4:
        lines.append(f'<a href="/media-fasilitas/list" target="_blank">Lihat semua media →</a>')
    lines.append(f"<br>📞 Hubungi: <b>{ADMIN_NO}</b>")
    return '<br>'.join(lines)


def _web_harga() -> str:
    xdata = _load_excel()
    lines = ["💰 <b>Harga Kamar Kost Kami</b><br>"]
    if xdata['harga']:
        for h in xdata['harga']:
            lines.append(f"🏠 <b>{h['tipe']}</b>")
            lines.append(f"&nbsp;&nbsp;• Bulanan: <b>{_fmt_rp(h['bulanan'])}</b>/bulan")
            if h['deposit']:
                lines.append(f"&nbsp;&nbsp;• Deposit: {_fmt_rp(h['deposit'])}")
            if h['ket']:
                lines.append(f"&nbsp;&nbsp;<i>{h['ket']}</i>")
            lines.append("")
    else:
        db_data = _kamar_dari_db()
        harga_set = sorted(set(v['harga'] for v in db_data.values() if v['harga'] > 0))
        if harga_set:
            for h in harga_set:
                lines.append(f"&nbsp;&nbsp;• {_fmt_rp(h)}/bulan")
        else:
            lines.append("Hubungi admin untuk info harga terkini.")
    lines.append(f"<br>📞 Info: <b>{ADMIN_NO}</b>")
    return '<br>'.join(lines)


def _web_fasilitas() -> str:
    xdata = _load_excel()
    lines = ["✨ <b>Fasilitas Kost Kami</b><br>"]
    if xdata['fasilitas']:
        by_kat: dict = {}
        for f in xdata['fasilitas']:
            kat = f['kategori'] or 'Umum'
            by_kat.setdefault(kat, []).append(f)
        for kat, items in by_kat.items():
            lines.append(f"<b>{kat}:</b>")
            for item in items:
                ico = '✅' if 'baik' in item['kondisi'].lower() else '⚠️'
                ket = f" <i>– {item['keterangan']}</i>" if item['keterangan'] else ''
                lines.append(f"&nbsp;&nbsp;{ico} {item['nama']}{ket}")
            lines.append("")
    else:
        lines.append(
            "✅ WiFi gratis<br>✅ Parkir motor & mobil<br>✅ Dapur bersama<br>"
            "✅ Mesin cuci<br>✅ CCTV 24 jam<br>✅ Air minum gratis"
        )
    lines.append(f"<br>📞 Info: <b>{ADMIN_NO}</b>")
    return '<br>'.join(lines)


def _web_admin() -> str:
    xdata = _load_excel()
    nama  = xdata['info'].get('Nama Admin / Pemilik', '')
    jam   = xdata['info'].get('Jam Layanan Admin', '08.00 – 21.00 WIB')
    no_hp = xdata['info'].get('No. HP / WA Admin', ADMIN_NO)
    return (
        f"📞 <b>Hubungi Admin</b><br><br>"
        f"{'<b>' + nama + '</b><br>' if nama else ''}"
        f"WA / Telp: <b>{no_hp}</b><br>"
        f"Jam: {jam}<br><br>"
        f"Silakan hubungi langsung ya kak 🙏"
    )


# ═════════════════════════════════════════════════════════════════════════════
# UTILITIES
# ═════════════════════════════════════════════════════════════════════════════

def _log_wa(sender: str, message: str, reply: str):
    try:
        conn   = get_db()
        suffix = sender[-9:]
        row    = conn.execute(
            "SELECT id FROM penghuni WHERE REPLACE(REPLACE(no_hp,' ',''),'-','') LIKE ? AND aktif=1",
            (f'%{suffix}',)
        ).fetchone()
        conn.execute(
            """INSERT INTO notif_wa (penghuni_id, pesan, status, tanggal_kirim)
               VALUES (?, ?, 'chatbot', datetime('now','localtime'))""",
            (row['id'] if row else None, f"[USER] {message}\n[BOT] {reply[:500]}")
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[WA Handler] Gagal log: {e}")


def _cleanup(state: dict, last_seen: dict):
    now     = time.time()
    expired = [k for k, ts in last_seen.items() if now - ts > _SESSION_TIMEOUT]
    for k in expired:
        state.pop(k, None)
        last_seen.pop(k, None)
