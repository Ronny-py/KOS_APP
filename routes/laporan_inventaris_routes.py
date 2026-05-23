"""
routes/laporan_inventaris_routes.py
Laporan inventaris barang kost — dengan export PDF & Excel.
"""
from flask import Blueprint, render_template, request, send_file
from utils.auth import login_required
from models.database import get_db
import io
from datetime import date

laporan_inventaris_bp = Blueprint('laporan_inventaris', __name__, url_prefix='/laporan/inventaris')

KONDISI_MAP = {
    'baik':         'Baik',
    'rusak':        'Rusak',
    'perlu_servis': 'Perlu Servis',
    'hilang':       'Hilang',
}

KATEGORI_MAP = {
    'elektronik':  'Elektronik',
    'furnitur':    'Furnitur',
    'kamar_mandi': 'Kamar Mandi',
    'dapur':       'Dapur',
    'keamanan':    'Keamanan',
    'kebersihan':  'Kebersihan',
    'lainnya':     'Lainnya',
}

KONDISI_MAP_ICON = {
    'baik':         '✅ Baik',
    'rusak':        '❌ Rusak',
    'perlu_servis': '⚠️ Perlu Servis',
    'hilang':       '🔍 Hilang',
}

KATEGORI_MAP_ICON = {
    'elektronik':  '💡 Elektronik',
    'furnitur':    '🪑 Furnitur',
    'kamar_mandi': '🚿 Kamar Mandi',
    'dapur':       '🍳 Dapur',
    'keamanan':    '🔒 Keamanan',
    'kebersihan':  '🧹 Kebersihan',
    'lainnya':     '📦 Lainnya',
}


def _get_data(filter_kat='', filter_kond='', filter_lokasi=''):
    conn = get_db()
    query  = "SELECT * FROM inventaris WHERE 1=1"
    params = []
    if filter_kat:
        query += " AND kategori=?";   params.append(filter_kat)
    if filter_kond:
        query += " AND kondisi=?";    params.append(filter_kond)
    if filter_lokasi:
        query += " AND lokasi LIKE ?"; params.append(f'%{filter_lokasi}%')
    query += " ORDER BY kategori, nama_barang"

    daftar    = conn.execute(query, params).fetchall()
    stats     = conn.execute("""
        SELECT COUNT(*) AS total_jenis, SUM(jumlah) AS total_unit,
               SUM(CASE WHEN kondisi='baik'         THEN jumlah ELSE 0 END) AS unit_baik,
               SUM(CASE WHEN kondisi='rusak'        THEN jumlah ELSE 0 END) AS unit_rusak,
               SUM(CASE WHEN kondisi='perlu_servis' THEN jumlah ELSE 0 END) AS unit_servis,
               SUM(CASE WHEN kondisi='hilang'       THEN jumlah ELSE 0 END) AS unit_hilang
        FROM inventaris
    """).fetchone()
    rekap_kat  = conn.execute("""
        SELECT kategori, COUNT(*) AS jenis, SUM(jumlah) AS total
        FROM inventaris GROUP BY kategori ORDER BY total DESC
    """).fetchall()
    rekap_kond = conn.execute("""
        SELECT kondisi, COUNT(*) AS jenis, SUM(jumlah) AS total
        FROM inventaris GROUP BY kondisi ORDER BY total DESC
    """).fetchall()
    lokasi_list = [r['lokasi'] for r in conn.execute(
        "SELECT DISTINCT lokasi FROM inventaris WHERE lokasi IS NOT NULL AND lokasi!='' ORDER BY lokasi"
    ).fetchall()]
    kat_list  = [r['kategori'] for r in conn.execute("SELECT DISTINCT kategori FROM inventaris ORDER BY kategori").fetchall()]
    kond_list = [r['kondisi']  for r in conn.execute("SELECT DISTINCT kondisi  FROM inventaris ORDER BY kondisi").fetchall()]
    conn.close()
    return daftar, stats, rekap_kat, rekap_kond, lokasi_list, kat_list, kond_list


# ── Halaman utama ─────────────────────────────────────────────────────────────
@laporan_inventaris_bp.route('/')
@login_required
def index():
    fk = request.args.get('kategori', '')
    fko = request.args.get('kondisi', '')
    fl = request.args.get('lokasi', '')
    daftar, stats, rekap_kat, rekap_kond, lokasi_list, kat_list, kond_list = _get_data(fk, fko, fl)
    return render_template('laporan/inventaris.html',
        daftar=daftar, stats=stats,
        rekap_kat=rekap_kat, rekap_kond=rekap_kond,
        lokasi_list=lokasi_list, kat_list=kat_list, kond_list=kond_list,
        filter_kat=fk, filter_kond=fko, filter_lokasi=fl,
        kondisi_map=KONDISI_MAP_ICON, kategori_map=KATEGORI_MAP_ICON,
    )


# ── Export Excel ──────────────────────────────────────────────────────────────
@laporan_inventaris_bp.route('/export/excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fk  = request.args.get('kategori', '')
    fko = request.args.get('kondisi', '')
    fl  = request.args.get('lokasi', '')
    daftar, stats, rekap_kat, rekap_kond, *_ = _get_data(fk, fko, fl)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaris"

    # Warna
    COLOR_HEADER = "1E3A5F"
    COLOR_SUBHEADER = "2E5984"
    COLOR_ACCENT = "4F6EF7"
    COLOR_LIGHT = "EEF2FF"
    COLOR_WHITE = "FFFFFF"
    COLOR_GRAY = "F5F5F5"

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(cell, text, bg=COLOR_HEADER, fg=COLOR_WHITE, bold=True, size=11, align='center'):
        cell.value = text
        cell.font = Font(bold=bold, color=fg, size=size, name='Arial')
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = border

    def data_cell(cell, text, bold=False, align='left', fg='000000'):
        cell.value = text
        cell.font = Font(name='Arial', size=10, bold=bold, color=fg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = border

    # ── Judul ──
    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = 'LAPORAN INVENTARIS BARANG'
    c.font = Font(bold=True, size=14, color=COLOR_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color=COLOR_HEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:H2')
    c = ws['A2']
    c.value = f'Tanggal Cetak: {date.today().strftime("%d %B %Y")}'
    c.font = Font(size=10, color='666666', name='Arial')
    c.alignment = Alignment(horizontal='center')
    c.fill = PatternFill('solid', start_color=COLOR_LIGHT)
    ws.row_dimensions[2].height = 18

    # ── Ringkasan statistik ──
    ws.row_dimensions[3].height = 8
    ws.merge_cells('A4:H4')
    c = ws['A4']
    c.value = 'RINGKASAN'
    c.font = Font(bold=True, size=11, color=COLOR_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color=COLOR_SUBHEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 22

    stat_labels = [
        ('Total Jenis', stats['total_jenis'] or 0),
        ('Total Unit', stats['total_unit'] or 0),
        ('Unit Baik', stats['unit_baik'] or 0),
        ('Perlu Servis', stats['unit_servis'] or 0),
        ('Rusak', stats['unit_rusak'] or 0),
        ('Hilang', stats['unit_hilang'] or 0),
    ]
    cols_stat = ['A','B','C','D','E','F']
    for i, (label, val) in enumerate(stat_labels):
        col = cols_stat[i]
        lc = ws[f'{col}5']
        lc.value = label
        lc.font = Font(bold=True, size=9, color='666666', name='Arial')
        lc.fill = PatternFill('solid', start_color=COLOR_GRAY)
        lc.alignment = Alignment(horizontal='center')
        lc.border = border
        vc = ws[f'{col}6']
        vc.value = val
        vc.font = Font(bold=True, size=16, color=COLOR_ACCENT, name='Arial')
        vc.fill = PatternFill('solid', start_color=COLOR_WHITE)
        vc.alignment = Alignment(horizontal='center')
        vc.border = border
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[7].height = 10

    # ── Header tabel ──
    headers = ['No', 'Nama Barang', 'Kategori', 'Jumlah', 'Kondisi', 'Lokasi', 'Keterangan', 'Tanggal']
    col_widths = [5, 30, 18, 10, 16, 16, 35, 14]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=8, column=i)
        hdr_cell(cell, h, bg=COLOR_SUBHEADER)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[8].height = 22

    # ── Data rows ──
    for r, row in enumerate(daftar, start=9):
        fill_bg = COLOR_LIGHT if r % 2 == 0 else COLOR_WHITE
        values = [
            r - 8,
            row['nama_barang'],
            KATEGORI_MAP.get(row['kategori'], row['kategori']),
            row['jumlah'],
            KONDISI_MAP.get(row['kondisi'], row['kondisi']),
            row['lokasi'] or '-',
            row['keterangan'] or '-',
            row['tanggal'] or '-',
        ]
        aligns = ['center','left','left','center','center','left','left','center']
        for c_idx, (val, aln) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=r, column=c_idx)
            data_cell(cell, val, align=aln)
            cell.fill = PatternFill('solid', start_color=fill_bg)
        ws.row_dimensions[r].height = 18

    # ── Total row ──
    total_row = 9 + len(daftar)
    ws.merge_cells(f'A{total_row}:C{total_row}')
    tc = ws[f'A{total_row}']
    hdr_cell(tc, 'TOTAL', bg=COLOR_HEADER, align='right')
    jml_cell = ws.cell(row=total_row, column=4)
    hdr_cell(jml_cell, f'=SUM(D9:D{total_row-1})', bg=COLOR_HEADER)
    for c_idx in range(5, 9):
        hdr_cell(ws.cell(row=total_row, column=c_idx), '', bg=COLOR_HEADER)
    ws.row_dimensions[total_row].height = 22

    # ── Sheet rekap per kategori ──
    ws2 = wb.create_sheet("Rekap Kategori")
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12

    ws2.merge_cells('A1:C1')
    c = ws2['A1']
    c.value = 'REKAP PER KATEGORI'
    c.font = Font(bold=True, size=13, color=COLOR_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color=COLOR_HEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    for i, h in enumerate(['Kategori', 'Jenis', 'Total Unit'], 1):
        cell = ws2.cell(row=2, column=i)
        hdr_cell(cell, h, bg=COLOR_SUBHEADER)

    for r, row in enumerate(rekap_kat, start=3):
        fill_bg = COLOR_LIGHT if r % 2 == 0 else COLOR_WHITE
        vals = [KATEGORI_MAP.get(row['kategori'], row['kategori']), row['jenis'], row['total']]
        alns = ['left', 'center', 'center']
        for ci, (v, a) in enumerate(zip(vals, alns), 1):
            cell = ws2.cell(row=r, column=ci)
            data_cell(cell, v, align=a)
            cell.fill = PatternFill('solid', start_color=fill_bg)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"inventaris_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


# ── Export PDF ────────────────────────────────────────────────────────────────
@laporan_inventaris_bp.route('/export/pdf')
@login_required
def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    fk  = request.args.get('kategori', '')
    fko = request.args.get('kondisi', '')
    fl  = request.args.get('lokasi', '')
    daftar, stats, rekap_kat, rekap_kond, *_ = _get_data(fk, fko, fl)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    C_DARK  = colors.HexColor('#1E3A5F')
    C_MID   = colors.HexColor('#2E5984')
    C_LIGHT = colors.HexColor('#EEF2FF')
    C_ACCENT= colors.HexColor('#4F6EF7')
    C_WHITE = colors.white
    C_GRAY  = colors.HexColor('#F5F5F5')

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', fontSize=16, fontName='Helvetica-Bold',
                                  textColor=C_DARK, alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle('sub',   fontSize=9,  fontName='Helvetica',
                                  textColor=colors.HexColor('#666666'), alignment=TA_CENTER, spaceAfter=12)
    sect_style  = ParagraphStyle('sect',  fontSize=11, fontName='Helvetica-Bold',
                                  textColor=C_DARK, spaceAfter=6, spaceBefore=10)

    story = []

    # Judul
    story.append(Paragraph('LAPORAN INVENTARIS BARANG', title_style))
    story.append(Paragraph(f'Tanggal Cetak: {date.today().strftime("%d %B %Y")}', sub_style))

    # Statistik ringkasan
    stat_data = [
        ['Total Jenis', 'Total Unit', 'Unit Baik', 'Perlu Servis', 'Rusak', 'Hilang'],
        [str(stats['total_jenis'] or 0), str(stats['total_unit'] or 0),
         str(stats['unit_baik'] or 0), str(stats['unit_servis'] or 0),
         str(stats['unit_rusak'] or 0), str(stats['unit_hilang'] or 0)],
    ]
    stat_tbl = Table(stat_data, colWidths=[40*mm]*6)
    stat_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), C_LIGHT),
        ('BACKGROUND', (0,1), (-1,1), C_WHITE),
        ('TEXTCOLOR',  (0,0), (-1,0), C_MID),
        ('TEXTCOLOR',  (0,1), (-1,1), C_ACCENT),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME',   (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 8),
        ('FONTSIZE',   (0,1), (-1,1), 16),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [C_LIGHT, C_WHITE]),
        ('BOX',        (0,0), (-1,-1), 0.5, C_MID),
        ('INNERGRID',  (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(stat_tbl)
    story.append(Spacer(1, 10))

    # Tabel utama
    story.append(Paragraph('Daftar Inventaris', sect_style))
    # Total lebar landscape A4 - margin = 267mm - 30mm = 237mm
    col_w = [8*mm, 45*mm, 28*mm, 12*mm, 25*mm, 25*mm, 70*mm, 24*mm]
    cell_style = ParagraphStyle('cell', fontSize=8, fontName='Helvetica',
                                 leading=10, wordWrap='LTR')
    tbl_data = [['No', 'Nama Barang', 'Kategori', 'Jml', 'Kondisi', 'Lokasi', 'Keterangan', 'Tanggal']]
    for i, row in enumerate(daftar, 1):
        tbl_data.append([
            str(i),
            Paragraph(row['nama_barang'] or '-', cell_style),
            KATEGORI_MAP.get(row['kategori'], row['kategori']),
            str(row['jumlah']),
            KONDISI_MAP.get(row['kondisi'], row['kondisi']),
            Paragraph(row['lokasi'] or '-', cell_style),
            Paragraph(row['keterangan'] or '-', cell_style),
            row['tanggal'] or '-',
        ])
    # Total
    total_unit = sum(row['jumlah'] for row in daftar)
    tbl_data.append(['', 'TOTAL', '', str(total_unit), '', '', '', ''])

    tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    row_colors = []
    for i in range(1, len(tbl_data)-1):
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        row_colors.append(('BACKGROUND', (0, i), (-1, i), bg))

    tbl.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0,0), (-1,0), C_DARK),
        ('TEXTCOLOR',  (0,0), (-1,0), C_WHITE),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 8),
        ('ALIGN',      (0,0), (-1,0), 'CENTER'),
        # Data
        ('FONTNAME',   (0,1), (-1,-2), 'Helvetica'),
        ('FONTSIZE',   (0,1), (-1,-2), 8),
        ('ALIGN',      (0,1), (0,-1), 'CENTER'),
        ('ALIGN',      (3,1), (3,-1), 'CENTER'),
        ('ALIGN',      (4,1), (4,-1), 'CENTER'),
        ('ALIGN',      (7,1), (7,-1), 'CENTER'),
        # Total row
        ('BACKGROUND', (0,-1), (-1,-1), C_MID),
        ('TEXTCOLOR',  (0,-1), (-1,-1), C_WHITE),
        ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN',      (0,-1), (-1,-1), 'CENTER'),
        # Grid
        ('BOX',        (0,0), (-1,-1), 0.5, C_MID),
        ('INNERGRID',  (0,0), (-1,-1), 0.3, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 4),
        ('RIGHTPADDING',  (0,0), (-1,-1), 4),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        *row_colors,
    ]))
    story.append(tbl)

    # Rekap per kategori
    story.append(Spacer(1, 12))
    story.append(Paragraph('Rekap per Kategori', sect_style))
    rek_data = [['Kategori', 'Jenis', 'Total Unit']]
    for row in rekap_kat:
        rek_data.append([
            KATEGORI_MAP.get(row['kategori'], row['kategori']),
            str(row['jenis']),
            str(row['total']),
        ])
    rek_tbl = Table(rek_data, colWidths=[80*mm, 40*mm, 40*mm])
    rek_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), C_MID),
        ('TEXTCOLOR',  (0,0), (-1,0), C_WHITE),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME',   (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('ALIGN',      (1,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [C_WHITE, C_LIGHT]),
        ('BOX',        (0,0), (-1,-1), 0.5, C_MID),
        ('INNERGRID',  (0,0), (-1,-1), 0.3, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(rek_tbl)

    doc.build(story)
    buf.seek(0)
    fname = f"inventaris_{date.today().strftime('%Y%m%d')}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=fname)
