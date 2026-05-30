"""
routes/komplain_routes.py
Admin: lihat, tanggapi, dan buat laporan komplain.
"""
import os
import requests as req_lib
from datetime import datetime
from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, send_from_directory, current_app, session)
from models import komplain_model
from functools import wraps

WA_SERVER_URL = "http://localhost:3000"

STATUS_LABEL = {
    'baru':      'Baru',
    'diproses':  'Sedang Diproses',
    'selesai':   'Selesai ✅',
    'ditolak':   'Ditolak ❌',
}

def _kirim_notif_wa_komplain(row, status_baru, catatan):
    """Kirim WA ke penghuni saat admin merespons komplain."""
    if not isinstance(row, dict):
        row = dict(row)  # sqlite3.Row → dict
    no_hp = (row.get('no_hp') or '').strip()
    if not no_hp:
        return

    # Format nomor: pastikan diawali 62
    nomor = no_hp.lstrip('+').strip()
    if nomor.startswith('0'):
        nomor = '62' + nomor[1:]

    status_text = STATUS_LABEL.get(status_baru, status_baru.upper())

    pesan = (
        f"📋 *Update Komplain Anda*\n\n"
        f"Halo *{row['nama_pelapor']}*, komplain Anda telah diperbarui.\n\n"
        f"📌 *Judul:* {row['judul']}\n"
        f"🏠 *Kamar:* {row['nomor_kamar']}\n"
        f"📊 *Status:* {status_text}\n"
    )
    if catatan:
        pesan += f"\n💬 *Catatan Admin:*\n{catatan}\n"

    pesan += "\nTerima kasih atas kesabaran Anda. 🙏"

    try:
        resp = req_lib.post(
            f"{WA_SERVER_URL}/api/send-message",
            json={"number": nomor, "message": pesan},
            timeout=5,
        )
        if resp.ok and resp.json().get('success'):
            current_app.logger.info(f"Notif WA komplain terkirim ke {nomor}")
        else:
            current_app.logger.warning(f"WA gagal: {resp.text}")
    except Exception as e:
        current_app.logger.warning(f"WA server tidak terjangkau: {e}")

komplain_bp = Blueprint('komplain', __name__, url_prefix='/komplain')


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_id'):
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


# ── Daftar Komplain ───────────────────────────────────────────────────────────
@komplain_bp.route('/')
@login_required
def index():
    status = request.args.get('status', '')
    search = request.args.get('q', '')
    now    = datetime.now()
    bulan  = int(request.args.get('bulan', now.month))
    tahun  = int(request.args.get('tahun', now.year))

    daftar = komplain_model.get_all(
        status=status or None,
        bulan=bulan if request.args.get('bulan') else None,
        tahun=tahun if request.args.get('tahun') else None,
        search=search or None,
    )

    # Hitung badge tiap status
    semua  = komplain_model.get_all()
    counts = {s: sum(1 for r in semua if r['status']==s)
              for s,_ in komplain_model.STATUS_LIST}
    counts['semua'] = len(semua)

    # Untuk tiap komplain di daftar, cari tanggal terakhir selesai
    # untuk kamar + kategori yang sama (tidak termasuk dirinya sendiri)
    from models.database import get_db
    terakhir_selesai = {}
    for r in daftar:
        row_ts = get_db().execute(
            """SELECT selesai_at, updated_at FROM komplain
               WHERE UPPER(TRIM(nomor_kamar)) = UPPER(TRIM(?))
                 AND LOWER(kategori)          = LOWER(?)
                 AND status                   = 'selesai'
                 AND id                       != ?
               ORDER BY COALESCE(selesai_at, updated_at) DESC
               LIMIT 1""",
            (r['nomor_kamar'], r['kategori'], r['id'])
        ).fetchone()
        if row_ts:
            tgl = row_ts['selesai_at'] or row_ts['updated_at']
            if tgl:
                terakhir_selesai[r['id']] = tgl[:10]

    return render_template('komplain/index.html',
        daftar=daftar,
        counts=counts,
        status_filter=status,
        search=search,
        bulan=bulan, tahun=tahun,
        STATUS_LIST=komplain_model.STATUS_LIST,
        KATEGORI_LIST=komplain_model.KATEGORI_LIST,
        PRIORITAS_LIST=komplain_model.PRIORITAS_LIST,
        terakhir_selesai=terakhir_selesai,
    )


# ── Tambah Komplain (Admin) ──────────────────────────────────────────────────
@komplain_bp.route('/tambah', methods=['POST'])
@login_required
def tambah():
    import os, uuid
    nama      = (request.form.get('nama')      or '').strip()
    kamar     = (request.form.get('kamar')     or '').strip()
    no_hp     = (request.form.get('no_hp')     or '').strip()
    kategori  = (request.form.get('kategori')  or 'lainnya').strip()
    judul     = (request.form.get('judul')      or '').strip()
    deskripsi = (request.form.get('deskripsi') or '').strip()
    prioritas = (request.form.get('prioritas') or 'normal').strip()

    if not nama or not kamar or not judul or not deskripsi:
        flash('Nama, nomor kamar, judul, dan deskripsi wajib diisi.', 'danger')
        return redirect(url_for('komplain.index'))

    # Simpan foto jika ada
    foto_path = None
    file = request.files.get('foto')
    if file and file.filename:
        ext = file.filename.rsplit('.', 1)[-1].lower()
        if ext in {'png', 'jpg', 'jpeg', 'webp', 'gif', 'heic'}:
            folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'komplain')
            os.makedirs(folder, exist_ok=True)
            fname = f"komplain_{uuid.uuid4().hex}.{ext}"
            file.save(os.path.join(folder, fname))
            foto_path = f"komplain/{fname}"

    komplain_model.tambah(nama, kamar, no_hp, kategori, judul, deskripsi, foto_path, prioritas)
    flash('Komplain berhasil ditambahkan.', 'success')
    return redirect(url_for('komplain.index'))


# ── Detail & Tanggapi ─────────────────────────────────────────────────────────
@komplain_bp.route('/<int:kid>')
@login_required
def detail(kid):
    row = komplain_model.get_by_id(kid)
    if not row:
        flash('Komplain tidak ditemukan.', 'danger')
        return redirect(url_for('komplain.index'))
    # Cari tanggal service AC terakhir jika kategori komplain ini adalah AC
    service_ac_terakhir = None
    if row['kategori'] == 'ac':
        from models.database import get_db
        ac_row = get_db().execute(
            """SELECT selesai_at, updated_at FROM komplain
               WHERE UPPER(TRIM(nomor_kamar)) = UPPER(TRIM(?))
                 AND LOWER(kategori) = 'ac'
                 AND status = 'selesai'
                 AND id != ?
               ORDER BY COALESCE(selesai_at, updated_at) DESC
               LIMIT 1""",
            (row['nomor_kamar'], kid)
        ).fetchone()
        if ac_row:
            tanggal_raw = ac_row['selesai_at'] or ac_row['updated_at']
            service_ac_terakhir = tanggal_raw[:10] if tanggal_raw else None

    return render_template('komplain/detail.html',
        row=row,
        STATUS_LIST=komplain_model.STATUS_LIST,
        PRIORITAS_LIST=komplain_model.PRIORITAS_LIST,
        KATEGORI_LIST=komplain_model.KATEGORI_LIST,
        service_ac_terakhir=service_ac_terakhir,
    )


@komplain_bp.route('/<int:kid>/tanggapi', methods=['POST'])
@login_required
def tanggapi(kid):
    status_lama = None
    row = komplain_model.get_by_id(kid)
    if row:
        status_lama = row['status']

    status  = request.form.get('status', 'diproses')
    catatan = request.form.get('catatan_admin', '').strip()
    komplain_model.update_status(kid, status, catatan or None)

    # Kirim notif WA ke penghuni jika status berubah atau ada catatan baru
    if row and (status != status_lama or catatan):
        _kirim_notif_wa_komplain(row, status, catatan)
        flash('Tanggapan berhasil disimpan dan notifikasi WA dikirim ke penghuni.', 'success')
    else:
        flash('Tanggapan berhasil disimpan.', 'success')

    return redirect(url_for('komplain.detail', kid=kid))


@komplain_bp.route('/<int:kid>/prioritas', methods=['POST'])
@login_required
def set_prioritas(kid):
    prioritas = request.form.get('prioritas', 'normal')
    komplain_model.update_prioritas(kid, prioritas)
    return redirect(url_for('komplain.detail', kid=kid))


@komplain_bp.route('/<int:kid>/kategori', methods=['POST'])
@login_required
def set_kategori(kid):
    kategori = request.form.get('kategori', 'lainnya')
    from models.database import get_db
    get_db().execute(
        'UPDATE komplain SET kategori = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
        (kategori, kid)
    )
    get_db().commit()
    flash('Kategori berhasil diubah.', 'success')
    return redirect(url_for('komplain.detail', kid=kid))


@komplain_bp.route('/<int:kid>/hapus', methods=['POST'])
@login_required
def hapus(kid):
    foto = komplain_model.hapus(kid)
    if foto:
        try:
            os.remove(os.path.join(current_app.config['UPLOAD_FOLDER'], foto))
        except Exception:
            pass
    flash('Komplain dihapus.', 'success')
    return redirect(url_for('komplain.index'))


# ── Serve foto ────────────────────────────────────────────────────────────────
@komplain_bp.route('/foto/<path:filename>')
@login_required
def serve_foto(filename):
    folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'komplain')
    return send_from_directory(folder, filename)





# ── API Riwayat Kamar ─────────────────────────────────────────────────────────
@komplain_bp.route('/api/riwayat-kamar')
@login_required
def api_riwayat_kamar():
    """
    GET /komplain/api/riwayat-kamar?kamar=A1
    Mengembalikan semua riwayat komplain untuk kamar tertentu.
    """
    from flask import jsonify
    from models.database import get_db
    kamar = (request.args.get('kamar') or '').strip()
    if not kamar:
        return jsonify(riwayat=[])

    KATEGORI_DICT = dict(komplain_model.KATEGORI_LIST)

    rows = get_db().execute(
        """SELECT id, judul, kategori, status, prioritas,
                  catatan_admin, created_at, selesai_at, updated_at
           FROM   komplain
           WHERE  UPPER(TRIM(nomor_kamar)) = UPPER(TRIM(?))
           ORDER  BY created_at DESC""",
        (kamar,)
    ).fetchall()

    riwayat = []
    for r in rows:
        selesai_raw = r['selesai_at'] or None
        riwayat.append({
            'id':            r['id'],
            'judul':         r['judul'],
            'kategori':      r['kategori'],
            'kategori_label': KATEGORI_DICT.get(r['kategori'], r['kategori']),
            'status':        r['status'],
            'prioritas':     r['prioritas'] or 'normal',
            'catatan_admin': r['catatan_admin'] or '',
            'created_at':    r['created_at'][:10] if r['created_at'] else '',
            'selesai_at':    selesai_raw[:10] if selesai_raw else '',
        })

    return jsonify(riwayat=riwayat)

# ── Export Excel ──────────────────────────────────────────────────────────────
@komplain_bp.route('/laporan/export-excel')
@login_required
def export_excel():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    bulan = int(request.args.get('bulan', datetime.now().month))
    tahun = int(request.args.get('tahun', datetime.now().year))

    BULAN_NAMA = ['','Januari','Februari','Maret','April','Mei','Juni',
                  'Juli','Agustus','September','Oktober','November','Desember']

    daftar  = komplain_model.get_all(bulan=bulan, tahun=tahun)
    _stats  = komplain_model.stats_bulan(bulan, tahun)
    stats   = dict(_stats) if _stats else {}
    per_kat = komplain_model.stats_per_kategori(bulan, tahun)
    avg     = komplain_model.avg_selesai_hari(bulan, tahun)

    KATEGORI_DICT = dict(komplain_model.KATEGORI_LIST)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Ringkasan ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ringkasan"

    HDR_FILL   = PatternFill("solid", fgColor="2F4FCC")
    HDR_FONT   = Font(color="FFFFFF", bold=True, size=11)
    TITLE_FONT = Font(bold=True, size=14, color="1A1A2E")
    SUB_FONT   = Font(bold=True, size=10, color="555555")
    BORDER     = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    CENTER = Alignment(horizontal='center', vertical='center')

    # Judul
    ws1.merge_cells('A1:D1')
    ws1['A1'] = f"LAPORAN KOMPLAIN & PERBAIKAN — {BULAN_NAMA[bulan].upper()} {tahun}"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = CENTER
    ws1.row_dimensions[1].height = 28

    ws1.append([])
    ws1.append(['Metrik', 'Jumlah'])
    for cell in ws1[3]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    rows_stat = [
        ('Total Komplain',   stats.get('total',    0)),
        ('Baru',             stats.get('baru',     0)),
        ('Sedang Diproses',  stats.get('diproses', 0)),
        ('Selesai',          stats.get('selesai',  0)),
        ('Ditolak',          stats.get('ditolak',  0)),
        ('Urgent',           stats.get('urgent',   0)),
        ('Rata-rata Selesai', f"{avg} hari" if avg else '—'),
    ]
    for label, val in rows_stat:
        ws1.append([label, val])
        for cell in ws1[ws1.max_row]:
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center')

    ws1.append([])
    ws1.append(['Kategori', 'Jumlah'])
    for cell in ws1[ws1.max_row]:
        cell.font = HDR_FONT
        cell.fill = PatternFill("solid", fgColor="1D6F42")
        cell.alignment = CENTER
        cell.border = BORDER

    for r in per_kat:
        label = KATEGORI_DICT.get(r['kategori'], r['kategori'])
        ws1.append([label, r['jumlah']])
        for cell in ws1[ws1.max_row]:
            cell.border = BORDER

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 16

    # ── Sheet 2: Detail ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detail Komplain")

    headers = ['#','Tanggal','Kamar','Pelapor','No. HP','Kategori',
               'Judul','Prioritas','Status','Selesai','Catatan Admin']
    ws2.append(headers)
    for i, cell in enumerate(ws2[1], 1):
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    STATUS_COLOR = {
        'selesai':  'C6EFCE',
        'diproses': 'FFEB9C',
        'baru':     'DDEEFF',
        'ditolak':  'FFC7CE',
    }

    for r in daftar:
        row = [
            r['id'],
            r['created_at'][:10],
            r['nomor_kamar'],
            r['nama_pelapor'],
            r['no_hp'] or '',
            KATEGORI_DICT.get(r['kategori'], r['kategori']),
            r['judul'],
            (r['prioritas'] or '').upper(),
            (r['status'] or '').upper(),
            r['selesai_at'][:10] if r['selesai_at'] else '—',
            r['catatan_admin'] or '—',
        ]
        ws2.append(row)
        fill_color = STATUS_COLOR.get(r['status'])
        for j, cell in enumerate(ws2[ws2.max_row], 1):
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=(j == 11))
            if fill_color:
                cell.fill = PatternFill("solid", fgColor=fill_color)

    col_widths = [6, 12, 10, 20, 16, 18, 35, 12, 12, 12, 30]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = ws2.dimensions

    # Output
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"laporan_komplain_{BULAN_NAMA[bulan]}_{tahun}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Export PDF ────────────────────────────────────────────────────────────────
@komplain_bp.route('/laporan/export-pdf')
@login_required
def export_pdf():
    import io
    from flask import send_file
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from PIL import Image as PILImage

    bulan = int(request.args.get('bulan', datetime.now().month))
    tahun = int(request.args.get('tahun', datetime.now().year))

    BULAN_NAMA = ['','Januari','Februari','Maret','April','Mei','Juni',
                  'Juli','Agustus','September','Oktober','November','Desember']

    daftar  = komplain_model.get_all(bulan=bulan, tahun=tahun)
    _stats  = komplain_model.stats_bulan(bulan, tahun)
    stats   = dict(_stats) if _stats else {}
    per_kat = komplain_model.stats_per_kategori(bulan, tahun)
    avg     = komplain_model.avg_selesai_hari(bulan, tahun)
    KATEGORI_DICT = dict(komplain_model.KATEGORI_LIST)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    ACCENT = colors.HexColor('#2F4FCC')
    GREEN  = colors.HexColor('#1D6F42')
    RED    = colors.HexColor('#C0392B')
    ORANGE = colors.HexColor('#E67E22')
    LIGHT  = colors.HexColor('#F0F4FF')
    GRAY   = colors.HexColor('#6B7280')

    style_title = ParagraphStyle('title', fontSize=18, fontName='Helvetica-Bold',
                                  textColor=ACCENT, alignment=TA_CENTER, spaceAfter=4)
    style_sub   = ParagraphStyle('sub', fontSize=10, fontName='Helvetica',
                                  textColor=GRAY, alignment=TA_CENTER, spaceAfter=12)
    style_h2    = ParagraphStyle('h2', fontSize=11, fontName='Helvetica-Bold',
                                  textColor=ACCENT, spaceBefore=14, spaceAfter=6)
    style_small = ParagraphStyle('small', fontSize=8, fontName='Helvetica',
                                  textColor=GRAY)

    story = []

    # ── Header ────────────────────────────────────────────────────────────
    story.append(Paragraph("LAPORAN KOMPLAIN & PERBAIKAN", style_title))
    story.append(Paragraph(f"{BULAN_NAMA[bulan]} {tahun}", style_sub))
    story.append(HRFlowable(width="100%", thickness=2, color=ACCENT, spaceAfter=14))

    # ── Ringkasan statistik ───────────────────────────────────────────────
    story.append(Paragraph("Ringkasan", style_h2))

    stat_data = [
        ['Total', 'Baru', 'Diproses', 'Selesai', 'Ditolak', 'Urgent', 'Avg Selesai'],
        [
            str(stats.get('total',    0)),
            str(stats.get('baru',     0)),
            str(stats.get('diproses', 0)),
            str(stats.get('selesai',  0)),
            str(stats.get('ditolak',  0)),
            str(stats.get('urgent',   0)),
            f"{avg} hari" if avg else '—',
        ],
    ]
    col_w = [3.5*cm] * 7
    stat_tbl = Table(stat_data, colWidths=col_w, rowHeights=[1*cm, 1.2*cm])
    stat_tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0), ACCENT),
        ('TEXTCOLOR',    (0,0), (-1,0), colors.white),
        ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,0), 9),
        ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME',     (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,1), (-1,1), 14),
        ('BACKGROUND',   (0,1), (-1,1), LIGHT),
        ('ROWBACKGROUNDS', (0,1), (-1,1), [LIGHT]),
        ('GRID',         (0,0), (-1,-1), 0.5, colors.white),
        ('ROUNDEDCORNERS', [4]),
    ]))
    story.append(stat_tbl)
    story.append(Spacer(1, 10))

    # ── Per kategori ──────────────────────────────────────────────────────
    if per_kat:
        story.append(Paragraph("Per Kategori", style_h2))
        kat_data = [['Kategori', 'Jumlah']]
        for r in per_kat:
            kat_data.append([KATEGORI_DICT.get(r['kategori'], r['kategori']), str(r['jumlah'])])
        kat_tbl = Table(kat_data, colWidths=[8*cm, 3*cm])
        kat_tbl.setStyle(TableStyle([
            ('BACKGROUND',  (0,0), (-1,0), GREEN),
            ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
            ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',    (0,0), (-1,-1), 9),
            ('ALIGN',       (1,0), (1,-1), 'CENTER'),
            ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAF9')]),
            ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
            ('ROWHEIGHT',   (0,0), (-1,-1), 0.7*cm),
        ]))
        story.append(kat_tbl)
        story.append(Spacer(1, 10))

    # ── Tabel detail ──────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#DDDDDD'), spaceAfter=8))
    story.append(Paragraph(f"Rincian Komplain — {len(daftar)} data", style_h2))

    tbl_data = [['#', 'Tanggal', 'Kamar', 'Pelapor', 'Kategori', 'Judul', 'Prioritas', 'Status', 'Selesai', 'Foto']]
    STATUS_COLOR_PDF = {
        'selesai':  colors.HexColor('#C6EFCE'),
        'diproses': colors.HexColor('#FFEB9C'),
        'baru':     colors.HexColor('#DDEEFF'),
        'ditolak':  colors.HexColor('#FFC7CE'),
    }
    PRIO_COLOR = {
        'urgent': RED,
        'tinggi': ORANGE,
        'normal': ACCENT,
        'rendah': GRAY,
    }

    from reportlab.platypus import Image as RLImage

    UPLOAD_FOLDER = current_app.config['UPLOAD_FOLDER']
    FOTO_SIZE     = 3 * cm   # tinggi & lebar thumbnail di tabel

    def _make_thumb(foto_path):
        """Buat RLImage thumbnail dari foto_path relatif ke UPLOAD_FOLDER. Return None jika gagal."""
        if not foto_path:
            return None
        full = os.path.join(UPLOAD_FOLDER, foto_path)
        if not os.path.isfile(full):
            return None
        try:
            with PILImage.open(full) as img:
                img.thumbnail((200, 200))
                tmp = io.BytesIO()
                fmt = img.format or 'JPEG'
                if fmt.upper() == 'HEIC':
                    fmt = 'JPEG'
                img.save(tmp, format=fmt)
                tmp.seek(0)
            return RLImage(tmp, width=FOTO_SIZE, height=FOTO_SIZE, kind='bound')
        except Exception as e:
            current_app.logger.warning(f"Gagal load foto {full}: {e}")
            return None

    row_heights = [0.8 * cm]   # header
    for r in daftar:
        thumb = _make_thumb(r['foto_path'] if 'foto_path' in r.keys() else None)
        row_h = FOTO_SIZE + 0.3 * cm if thumb else 0.75 * cm
        row_heights.append(row_h)

        style_judul = ParagraphStyle('j', fontSize=8, leading=10)
        tbl_data.append([
            str(r['id']),
            r['created_at'][:10],
            r['nomor_kamar'],
            r['nama_pelapor'],
            KATEGORI_DICT.get(r['kategori'], r['kategori']),
            Paragraph(r['judul'], style_judul),
            (r['prioritas'] or '').upper(),
            (r['status']    or '').upper(),
            r['selesai_at'][:10] if r['selesai_at'] else '—',
            thumb if thumb else Paragraph('—', ParagraphStyle('nd', fontSize=8, textColor=GRAY, alignment=TA_CENTER)),
        ])

    col_w2 = [1*cm, 2.2*cm, 1.8*cm, 3.2*cm, 3*cm, 5.5*cm, 2.2*cm, 2.2*cm, 2.2*cm, FOTO_SIZE + 0.2*cm]
    det_tbl = Table(tbl_data, colWidths=col_w2, rowHeights=row_heights, repeatRows=1)

    tbl_style = [
        ('BACKGROUND',  (0,0), (-1,0), ACCENT),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,-1), 8),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('ALIGN',       (5,1), (5,-1), 'LEFT'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',        (0,0), (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
    ]
    # Warna baris per status
    for i, r in enumerate(daftar, 1):
        c = STATUS_COLOR_PDF.get(r['status'])
        if c:
            tbl_style.append(('BACKGROUND', (0,i), (-1,i), c))

    det_tbl.setStyle(TableStyle(tbl_style))
    story.append(det_tbl)

    # ── Halaman lampiran foto (hanya komplain yg punya foto) ──────────────
    punya_foto = [r for r in daftar
                  if ('foto_path' in r.keys()) and r['foto_path']
                  and os.path.isfile(os.path.join(UPLOAD_FOLDER, r['foto_path']))]

    if punya_foto:
        from reportlab.platypus import PageBreak, KeepTogether
        story.append(PageBreak())
        story.append(Paragraph("Lampiran Foto Bukti Komplain", style_title))
        story.append(Spacer(1, 6))
        story.append(HRFlowable(width="100%", thickness=2, color=ACCENT, spaceAfter=16))

        FOTO_BESAR = 8 * cm
        style_label = ParagraphStyle('lbl', fontSize=9, fontName='Helvetica-Bold',
                                      textColor=ACCENT, spaceAfter=2)
        style_desc  = ParagraphStyle('dsc', fontSize=8, fontName='Helvetica',
                                      textColor=GRAY, spaceAfter=10)

        # Susun 2 kolom per baris
        pairs = [punya_foto[i:i+2] for i in range(0, len(punya_foto), 2)]
        for pair in pairs:
            cells = []
            for r in pair:
                full = os.path.join(UPLOAD_FOLDER, r['foto_path'])
                try:
                    with PILImage.open(full) as img:
                        iw, ih = img.size
                        ratio   = min(FOTO_BESAR / iw, FOTO_BESAR / ih)
                        disp_w  = iw * ratio
                        disp_h  = ih * ratio
                        tmp = io.BytesIO()
                        fmt = img.format or 'JPEG'
                        if fmt.upper() == 'HEIC':
                            fmt = 'JPEG'
                        img.save(tmp, format=fmt)
                        tmp.seek(0)
                    foto_el = RLImage(tmp, width=disp_w, height=disp_h)
                except Exception as e:
                    current_app.logger.warning(f"Lampiran foto gagal: {e}")
                    foto_el = Paragraph('[Foto tidak dapat dimuat]',
                                        ParagraphStyle('err', fontSize=8, textColor=RED))

                kat_label = KATEGORI_DICT.get(r['kategori'], r['kategori'])
                label_el  = Paragraph(
                    f"#{r['id']} — Kamar {r['nomor_kamar']} ({r['nama_pelapor']})",
                    style_label)
                desc_el   = Paragraph(
                    f"{kat_label} · {(r['prioritas'] or '').upper()} · {(r['status'] or '').upper()} · {r['created_at'][:10]}",
                    style_desc)
                cells.append([label_el, desc_el, foto_el])

            # Pad ke 2 kolom
            while len(cells) < 2:
                cells.append(['', '', ''])

            foto_tbl = Table(
                [[cells[0][0], '', cells[1][0]],
                 [cells[0][1], '', cells[1][1]],
                 [cells[0][2], '', cells[1][2]]],
                colWidths=[12*cm, 1*cm, 12*cm],
            )
            foto_tbl.setStyle(TableStyle([
                ('VALIGN',  (0,0), (-1,-1), 'TOP'),
                ('ALIGN',   (0,0), (-1,-1), 'LEFT'),
                ('TOPPADDING',    (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(KeepTogether([foto_tbl, Spacer(1, 14)]))

    # Footer
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"Dicetak: {datetime.now().strftime('%d %B %Y %H:%M')}  |  Sistem Manajemen Kost",
        style_small
    ))

    doc.build(story)
    buf.seek(0)
    fname = f"laporan_komplain_{BULAN_NAMA[bulan]}_{tahun}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')



@komplain_bp.route('/laporan')
@login_required
def laporan():
    now   = datetime.now()
    bulan = int(request.args.get('bulan', now.month))
    tahun = int(request.args.get('tahun', now.year))

    stats     = komplain_model.stats_bulan(bulan, tahun)
    per_kat   = komplain_model.stats_per_kategori(bulan, tahun)
    avg_hari  = komplain_model.avg_selesai_hari(bulan, tahun)
    daftar    = komplain_model.get_all(bulan=bulan, tahun=tahun)

    BULAN_NAMA = ['','Januari','Februari','Maret','April','Mei','Juni',
                  'Juli','Agustus','September','Oktober','November','Desember']

    return render_template('komplain/laporan.html',
        bulan=bulan, tahun=tahun,
        bulan_nama=BULAN_NAMA[bulan],
        stats=stats,
        per_kat=per_kat,
        avg_hari=avg_hari,
        daftar=daftar,
        KATEGORI_LIST=komplain_model.KATEGORI_LIST,
        STATUS_LIST=komplain_model.STATUS_LIST,
        BULAN_NAMA=BULAN_NAMA,
    )
